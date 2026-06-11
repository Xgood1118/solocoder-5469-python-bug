import os
from datetime import datetime

from openpyxl import Workbook
from config import EXPORTS_DIR


def export_predictions_to_excel(predictions, original_descriptions=None, output_filename=None):
    if output_filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"predictions_{timestamp}.xlsx"

    os.makedirs(EXPORTS_DIR, exist_ok=True)
    output_path = os.path.join(EXPORTS_DIR, output_filename)

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "预测汇总"
    headers = ["序号", "原始描述", "预测分类", "预测严重程度", "分类置信度", "严重程度置信度",
               "综合置信度", "需人工确认", "模型版本", "预测耗时(ms)"]
    ws_summary.append(headers)

    for i, pred in enumerate(predictions):
        desc = original_descriptions[i] if original_descriptions and i < len(original_descriptions) else ""
        row = [
            i + 1,
            desc,
            pred.get("category", ""),
            pred.get("severity", ""),
            pred.get("category_confidence", ""),
            pred.get("severity_confidence", ""),
            pred.get("confidence", ""),
            "是" if pred.get("needs_review") else "否",
            pred.get("model_version", ""),
            pred.get("predict_time_ms", ""),
        ]
        ws_summary.append(row)

    ws_review = wb.create_sheet("待人工确认")
    review_headers = ["序号", "原始描述", "预测分类", "预测严重程度", "确认分类", "确认严重程度", "确认人"]
    ws_review.append(review_headers)

    review_idx = 1
    for i, pred in enumerate(predictions):
        if pred.get("needs_review"):
            desc = original_descriptions[i] if original_descriptions and i < len(original_descriptions) else ""
            row = [
                review_idx,
                desc,
                pred.get("category", ""),
                pred.get("severity", ""),
                "",
                "",
                "",
            ]
            ws_review.append(row)
            review_idx += 1

    ws_final = wb.create_sheet("最终分类")
    final_headers = ["序号", "原始描述", "最终分类", "最终严重程度", "来源(模型/人工)", "确认人"]
    ws_final.append(final_headers)

    for i, pred in enumerate(predictions):
        desc = original_descriptions[i] if original_descriptions and i < len(original_descriptions) else ""
        source = "人工" if pred.get("confirmed") else "模型"
        row = [
            i + 1,
            desc,
            pred.get("category", ""),
            pred.get("severity", ""),
            source,
            pred.get("confirmer", ""),
        ]
        ws_final.append(row)

    wb.save(output_path)
    return output_path
